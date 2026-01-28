#!/usr/bin/env python3
"""
Phase 3: Report generation for Grade Processing.
Loads templates/template.xlsx, reads headers dynamically, fills data.

Key pattern:
  - Template is editable by non-technical users (teachers)
  - generate.py reads headers from template to determine column mapping
  - Formatting (colors, borders, fonts) is preserved from template
  - Users can rename/add/remove columns and this script adapts
"""

import argparse
import json
import sys
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

import yaml


# =============================================================================
# Configuration
# =============================================================================

TEMPLATE_DIR = Path(__file__).parent.parent / "templates"
CONFIG_DIR = Path(__file__).parent.parent / "config"


def load_config() -> dict:
    """Load settings.yaml."""
    config_path = CONFIG_DIR / "settings.yaml"
    with open(config_path, encoding="utf-8") as f:
        return yaml.safe_load(f)


# =============================================================================
# Template Loading
# =============================================================================

def load_template(config: dict) -> tuple:
    """Load template workbook."""
    template_file = config["template"]["file"]
    template_path = TEMPLATE_DIR / template_file

    if not template_path.exists():
        print(f"[ERROR] Template not found: {template_path}")
        sys.exit(1)

    wb = load_workbook(template_path)
    print(f"  [TMPL] Loaded: {template_path}")
    print(f"  [TMPL] Sheets: {wb.sheetnames}")
    return wb


def parse_headers(ws, header_row: int) -> dict:
    """
    Dynamically read column headers from a worksheet row.
    Returns: {header_name: column_index}

    This is the key pattern that makes the template user-editable:
    - Teachers can rename columns → mapping updates automatically
    - Teachers can add columns → new columns are discovered
    - Teachers can reorder columns → positions are tracked correctly
    """
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            headers[str(val).strip()] = col
    return headers


def copy_cell_style(source_cell, target_cell):
    """Copy formatting from source cell to target cell."""
    if source_cell.font:
        target_cell.font = copy(source_cell.font)
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)
    if source_cell.border:
        target_cell.border = copy(source_cell.border)
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format


# =============================================================================
# Sheet Writers
# =============================================================================

# Column name → data key mapping (template header → JSON field)
REPORT_FIELD_MAP = {
    "학번": "student_id",
    "이름": "name",
    "중간고사": "midterm",
    "기말고사": "final",
    "과제평균": "assignment_avg",
    "출석률": "attendance_pct",
    "총점": "total_score",
    "등급": "grade",
    "코멘트": "comment",
}

AT_RISK_FIELD_MAP = {
    "학번": "student_id",
    "이름": "name",
    "총점": "total_score",
    "등급": "grade",
    "주요 문제": "issues_text",
    "권고 사항": "recommendation",
}

# Highlight colors
WARN_FONT = Font(color="CC0000")
HIGHLIGHT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def write_report_sheet(wb, config, results):
    """Write student grades to the 성적표 sheet."""
    sheet_name = config["template"]["sheets"]["report"]
    if sheet_name not in wb.sheetnames:
        print(f"  [WARN] Sheet '{sheet_name}' not found, skipping")
        return

    ws = wb[sheet_name]
    header_row = config["template"]["data_start_row"]["report"] - 1  # headers are 1 row above data
    data_start = config["template"]["data_start_row"]["report"]

    headers = parse_headers(ws, header_row)
    print(f"  [SHEET] {sheet_name}: headers={list(headers.keys())}")

    # Write metadata
    ws.cell(row=2, column=2, value="2024-2학기")
    ws.cell(row=2, column=5, value="종합")
    ws.cell(row=2, column=8, value="자동생성")

    # Use first data row as style reference
    style_ref_row = data_start

    for idx, student in enumerate(results):
        row = data_start + idx

        for header_name, col_idx in headers.items():
            field_key = REPORT_FIELD_MAP.get(header_name)
            if not field_key:
                continue

            value = student.get(field_key, "")
            cell = ws.cell(row=row, column=col_idx, value=value)

            # Copy style from template's pre-formatted row
            ref_cell = ws.cell(row=style_ref_row, column=col_idx)
            copy_cell_style(ref_cell, cell)

            # Highlight at-risk students
            if field_key == "total_score" and isinstance(value, (int, float)) and value < 70:
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color="CC0000")
            if field_key == "grade" and value == "F":
                cell.fill = HIGHLIGHT_FILL

    print(f"  [OK] {sheet_name}: {len(results)} students written")


def write_summary_sheet(wb, config, statistics):
    """Write class statistics to the 반 요약 sheet."""
    sheet_name = config["template"]["sheets"]["summary"]
    if sheet_name not in wb.sheetnames:
        print(f"  [WARN] Sheet '{sheet_name}' not found, skipping")
        return

    ws = wb[sheet_name]
    data_start = config["template"]["data_start_row"]["summary"]

    # Left section: key statistics
    summary_rows = [
        ("수강 인원", f"{statistics['student_count']}명", ""),
        ("평균 점수", f"{statistics['score_avg']}점", ""),
        ("최고 점수", f"{statistics['score_max']}점", ""),
        ("최저 점수", f"{statistics['score_min']}점", ""),
        ("평균 출석률", f"{statistics['attendance_avg']}%", ""),
        ("관찰 대상", f"{statistics['at_risk_count']}명",
         "총점 70점 미만 또는 출석 80% 미만"),
    ]

    style_ref_row = data_start
    for idx, (label, value, note) in enumerate(summary_rows):
        row = data_start + idx
        for col, val in [(1, label), (2, value), (3, note)]:
            cell = ws.cell(row=row, column=col, value=val)
            ref_cell = ws.cell(row=style_ref_row, column=col)
            copy_cell_style(ref_cell, cell)

    # Right section: grade distribution
    grade_order = ["A+", "A", "B+", "B", "C+", "C", "D+", "D", "F"]
    grade_dist = statistics.get("grade_distribution", {})
    total = statistics["student_count"]

    for idx, grade in enumerate(grade_order):
        row = data_start + idx
        count = grade_dist.get(grade, 0)
        pct = f"{count / total * 100:.0f}%" if total > 0 else "0%"

        ws.cell(row=row, column=5, value=grade)
        ws.cell(row=row, column=6, value=count)
        ws.cell(row=row, column=7, value=pct)

        # Copy style from template
        for col in range(5, 8):
            ref_cell = ws.cell(row=style_ref_row, column=col)
            cell = ws.cell(row=row, column=col)
            copy_cell_style(ref_cell, cell)

    print(f"  [OK] {sheet_name}: statistics written")


def write_at_risk_sheet(wb, config, at_risk_students):
    """Write at-risk student list to the 관찰 대상 sheet."""
    sheet_name = config["template"]["sheets"]["at_risk"]
    if sheet_name not in wb.sheetnames:
        print(f"  [WARN] Sheet '{sheet_name}' not found, skipping")
        return

    ws = wb[sheet_name]
    header_row = config["template"]["data_start_row"]["at_risk"] - 1
    data_start = config["template"]["data_start_row"]["at_risk"]

    headers = parse_headers(ws, header_row)
    print(f"  [SHEET] {sheet_name}: headers={list(headers.keys())}")

    style_ref_row = data_start

    for idx, student in enumerate(at_risk_students):
        row = data_start + idx

        # Prepare derived fields
        student_data = {
            **student,
            "issues_text": ", ".join(student.get("issues", [])),
            "recommendation": student.get("recommendation", ""),
        }

        for header_name, col_idx in headers.items():
            field_key = AT_RISK_FIELD_MAP.get(header_name)
            if not field_key:
                continue

            value = student_data.get(field_key, "")
            cell = ws.cell(row=row, column=col_idx, value=value)

            ref_cell = ws.cell(row=style_ref_row, column=col_idx)
            copy_cell_style(ref_cell, cell)

    print(f"  [OK] {sheet_name}: {len(at_risk_students)} at-risk students written")


# =============================================================================
# Main
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Grade Processing - Report Generation (Phase 3)"
    )
    parser.add_argument(
        "--judged-data", required=True,
        help="LLM 판단 완료 데이터 (judged_data.json)",
    )
    parser.add_argument(
        "--output-dir", "-o", required=True,
        help="출력 디렉토리",
    )
    args = parser.parse_args()

    judged_path = Path(args.judged_data).resolve()
    if not judged_path.exists():
        print(f"[ERROR] File not found: {judged_path}")
        sys.exit(1)

    with open(judged_path, encoding="utf-8") as f:
        data = json.load(f)

    config = load_config()

    print(f"\n=== Report Generation (Excel Template) ===\n")

    # Load template
    wb = load_template(config)

    # Write all sheets
    write_report_sheet(wb, config, data.get("results", []))
    write_summary_sheet(wb, config, data.get("statistics", {}))
    write_at_risk_sheet(wb, config, data.get("at_risk", []))

    # Also save JSON report
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    json_path = output_dir / "grade_report.json"
    report_json = {
        "meta": {
            **data.get("meta", {}),
            "generated_at": datetime.now().isoformat(),
        },
        "results": data.get("results", []),
        "at_risk": data.get("at_risk", []),
        "statistics": data.get("statistics", {}),
    }
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(report_json, f, ensure_ascii=False, indent=2)
    print(f"\n  [OK] JSON: {json_path}")

    # Save Excel
    xlsx_path = output_dir / "grade_report.xlsx"
    wb.save(xlsx_path)
    print(f"  [OK] Excel: {xlsx_path}")

    print(f"\n[OK] Report generation complete: {output_dir}")


if __name__ == "__main__":
    main()
