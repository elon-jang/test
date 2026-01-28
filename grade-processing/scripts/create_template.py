#!/usr/bin/env python3
"""
Helper: Create the Excel template for grade-processing.
Run once to generate templates/template.xlsx.

The generated template is designed to be editable by non-technical users:
- Teachers can rename column headers
- Teachers can add/remove columns
- Teachers can change formatting (colors, fonts, borders)
- generate.py reads headers dynamically, so changes are automatically reflected
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).parent.parent / "templates" / "template.xlsx"

# Shared styles
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True)
DATA_FONT = Font(name="맑은 고딕", size=10)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
# Accent colors for alternating summary rows
LIGHT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def style_header_row(ws, row, col_start, col_end):
    """Apply header style to a range of cells."""
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def create_report_sheet(wb):
    """Sheet 1: 성적표 - Student grade report."""
    ws = wb.active
    ws.title = "성적표"

    # Row 1: Title
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "성적표"
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER

    # Row 2: Metadata placeholders
    ws["A2"] = "학기:"
    ws["B2"] = ""  # filled by generate.py
    ws["D2"] = "과목:"
    ws["E2"] = ""
    ws["G2"] = "담당교사:"
    ws["H2"] = ""
    for col in range(1, 10):
        ws.cell(row=2, column=col).font = Font(name="맑은 고딕", size=10)

    # Row 3: Empty separator

    # Row 4: Headers
    headers = ["학번", "이름", "중간고사", "기말고사", "과제평균", "출석률", "총점", "등급", "코멘트"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=4, column=col_idx, value=header)
    style_header_row(ws, 4, 1, len(headers))

    # Column widths
    widths = [10, 10, 10, 10, 10, 10, 8, 8, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Pre-format data rows (5-30) with borders and alignment
    for row in range(5, 31):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col <= 2:
                cell.alignment = LEFT
            elif col == 9:  # 코멘트
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = CENTER

    return ws


def create_summary_sheet(wb):
    """Sheet 2: 반 요약 - Class summary statistics."""
    ws = wb.create_sheet("반 요약")

    # Row 1: Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "반 요약"
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER

    # Row 2: Empty

    # Row 3: Headers
    headers = ["항목", "값", "비고", ""]
    for col_idx, header in enumerate(headers[:3], 1):
        ws.cell(row=3, column=col_idx, value=header)
    style_header_row(ws, 3, 1, 3)

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 30

    # Pre-format data rows
    for row in range(4, 25):
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col != 2 else CENTER

    # Row 6+: Grade distribution headers
    ws.cell(row=3, column=5, value="등급")
    ws.cell(row=3, column=6, value="인원")
    ws.cell(row=3, column=7, value="비율")
    style_header_row(ws, 3, 5, 7)
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 10

    for row in range(4, 14):
        for col in range(5, 8):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER

    return ws


def create_at_risk_sheet(wb):
    """Sheet 3: 관찰 대상 - At-risk student list."""
    ws = wb.create_sheet("관찰 대상")

    # Row 1: Title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "관찰 대상 학생"
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER

    # Row 2: Empty

    # Row 3: Headers
    headers = ["학번", "이름", "총점", "등급", "주요 문제", "권고 사항"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=3, column=col_idx, value=header)
    style_header_row(ws, 3, 1, len(headers))

    # Column widths
    widths = [10, 10, 8, 8, 35, 35]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Pre-format data rows with warning fill
    for row in range(4, 20):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col >= 5:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif col <= 2:
                cell.alignment = LEFT
            else:
                cell.alignment = CENTER

    return ws


def main():
    wb = Workbook()
    create_report_sheet(wb)
    create_summary_sheet(wb)
    create_at_risk_sheet(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"[OK] Template created: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
