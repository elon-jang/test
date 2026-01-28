#!/usr/bin/env python3
"""
Helper: Create the Excel template for invoice-classification.
Run once to generate templates/template.xlsx.

The template is designed for finance team editing:
- Column headers can be renamed
- Formatting (colors, borders) can be customized
- Sheets can be rearranged
- generate.py reads headers dynamically
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).parent.parent / "templates" / "template.xlsx"

# Styles
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="맑은 고딕", size=14, bold=True)
DATA_FONT = Font(name="맑은 고딕", size=10)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
AMOUNT_FORMAT = '#,##0"원"'
PCT_FORMAT = '0.0"%"'

# Status colors
STATUS_FILLS = {
    "APPROVED": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "PENDING": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "FLAG": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "REVIEW": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}


def style_header_row(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def create_summary_sheet(wb):
    """Sheet 1: 요약 - Category summary + status counts."""
    ws = wb.active
    ws.title = "요약"

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"].value = "경비보고서 요약"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    # Left: Category summary
    headers_left = ["카테고리", "GL코드", "건수", "합계", "비율"]
    for i, h in enumerate(headers_left, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 1, 5)

    widths_left = [12, 10, 8, 15, 10]
    for i, w in enumerate(widths_left, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Right: Status summary
    ws.cell(row=3, column=7, value="정책상태")
    ws.cell(row=3, column=8, value="건수")
    style_header_row(ws, 3, 7, 8)
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 8

    # Pre-format data rows
    for row in range(4, 15):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col == 4:
                cell.number_format = AMOUNT_FORMAT
                cell.alignment = RIGHT
            elif col == 5:
                cell.alignment = CENTER
            elif col >= 3:
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT
        for col in range(7, 9):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER

    return ws


def create_detail_sheet(wb):
    """Sheet 2: 상세 - All expense items."""
    ws = wb.create_sheet("상세")

    ws.merge_cells("A1:I1")
    ws["A1"].value = "경비 상세 내역"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    # Metadata row
    ws["A2"] = "기간:"
    ws["B2"] = ""
    ws["D2"] = "부서:"
    ws["E2"] = ""
    ws["G2"] = "신청자:"
    ws["H2"] = ""
    for col in range(1, 10):
        ws.cell(row=2, column=col).font = Font(name="맑은 고딕", size=10)

    headers = ["번호", "날짜", "거래처", "내역", "금액", "GL코드", "카테고리", "정책상태", "비고"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 1, len(headers))

    widths = [8, 12, 18, 25, 13, 10, 10, 10, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in range(4, 30):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col == 5:  # 금액
                cell.number_format = AMOUNT_FORMAT
                cell.alignment = RIGHT
            elif col in (1, 2, 6, 7, 8):
                cell.alignment = CENTER
            elif col == 9:  # 비고
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = LEFT

    return ws


def create_flag_sheet(wb):
    """Sheet 3: 플래그 - Flagged/pending items."""
    ws = wb.create_sheet("플래그")

    ws.merge_cells("A1:H1")
    ws["A1"].value = "플래그 항목 (검토 필요)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER

    headers = ["번호", "날짜", "거래처", "내역", "금액", "상태", "사유", "조치"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 1, len(headers))

    widths = [8, 12, 18, 25, 13, 10, 35, 35]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in range(4, 20):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col == 5:
                cell.number_format = AMOUNT_FORMAT
                cell.alignment = RIGHT
            elif col in (1, 2, 6):
                cell.alignment = CENTER
            elif col >= 7:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = LEFT

    return ws


def main():
    wb = Workbook()
    create_summary_sheet(wb)
    create_detail_sheet(wb)
    create_flag_sheet(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"[OK] Template created: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
