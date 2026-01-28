#!/usr/bin/env python3
"""
Phase 3: Report generation for Invoice Classification.
Loads templates/template.xlsx, reads headers dynamically, fills data.

★ Excel Template Pattern:
  - Template editable by finance team (non-technical users)
  - parse_headers() reads column positions dynamically
  - copy_cell_style() preserves formatting from template
  - Multi-sheet output: 요약 + 상세 + 플래그
"""

import argparse
import json
import sys
from collections import defaultdict
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

import yaml


# =============================================================================
# Configuration
# =============================================================================

TEMPLATE_DIR = Path(__file__).parent.parent / "templates"
CONFIG_DIR = Path(__file__).parent.parent / "config"

SHEET_CONFIG = {
    "summary": {"name": "요약", "data_start": 4, "header_row": 3},
    "detail": {"name": "상세", "data_start": 4, "header_row": 3},
    "flag": {"name": "플래그", "data_start": 4, "header_row": 3},
}

# Column mapping: template header → JSON field
DETAIL_FIELD_MAP = {
    "번호": "id",
    "날짜": "date",
    "거래처": "vendor",
    "내역": "description",
    "금액": "amount",
    "GL코드": "gl_code",
    "카테고리": "category",
    "정책상태": "policy_status",
    "비고": "rationale",
}

FLAG_FIELD_MAP = {
    "번호": "id",
    "날짜": "date",
    "거래처": "vendor",
    "내역": "description",
    "금액": "amount",
    "상태": "policy_status",
    "사유": "flag",
    "조치": "rationale",
}

# Status formatting
STATUS_FILLS = {
    "APPROVED": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "PENDING": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "FLAG": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "REVIEW": PatternFill(start_color="E6B8AF", end_color="E6B8AF", fill_type="solid"),
}

STATUS_LABELS = {
    "APPROVED": "승인",
    "PENDING": "승인대기",
    "FLAG": "플래그",
    "REVIEW": "검토필요",
}


def load_config() -> dict:
    with open(CONFIG_DIR / "settings.yaml", encoding="utf-8") as f:
        return yaml.safe_load(f)


# =============================================================================
# Template Helpers
# =============================================================================

def load_template(config: dict):
    template_path = TEMPLATE_DIR / "template.xlsx"
    if not template_path.exists():
        print(f"[ERROR] Template not found: {template_path}")
        sys.exit(1)
    wb = load_workbook(template_path)
    print(f"  [TMPL] Loaded: {template_path}")
    print(f"  [TMPL] Sheets: {wb.sheetnames}")
    return wb


def parse_headers(ws, header_row: int) -> dict:
    """Dynamically read column headers → {name: col_index}."""
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            headers[str(val).strip()] = col
    return headers


def copy_cell_style(source_cell, target_cell):
    if source_cell.font:
        target_cell.font = copy(source_cell.font)
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)
    if source_cell.border:
        target_cell.border = copy(source_cell.border)
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format


# =============================================================================
# Sheet Writers
# =============================================================================

def write_summary_sheet(wb, items, summary):
    """Write category summary + status counts to 요약 sheet."""
    cfg = SHEET_CONFIG["summary"]
    ws = wb[cfg["name"]]
    data_start = cfg["data_start"]
    style_ref = data_start

    # Left section: by category
    by_cat = summary.get("by_category", {})
    total_amount = sum(info["total"] for info in by_cat.values()) if by_cat else 0
    row = data_start

    for cat, info in by_cat.items():
        pct = f"{info['total'] / total_amount * 100:.1f}%" if total_amount > 0 else "0%"
        values = [cat, info.get("gl_code", ""), info["count"], info["total"], pct]
        for col_offset, val in enumerate(values):
            cell = ws.cell(row=row, column=col_offset + 1, value=val)
            ref = ws.cell(row=style_ref, column=col_offset + 1)
            copy_cell_style(ref, cell)
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="합계").font = Font(name="맑은 고딕", size=10, bold=True)
    ws.cell(row=row, column=3, value=len(items))
    cell_total = ws.cell(row=row, column=4, value=total_amount)
    ref = ws.cell(row=style_ref, column=4)
    copy_cell_style(ref, cell_total)
    cell_total.font = Font(name="맑은 고딕", size=10, bold=True)
    ws.cell(row=row, column=5, value="100%")

    # Right section: by status
    by_status = summary.get("by_status", {})
    row = data_start
    for status, count in by_status.items():
        label = STATUS_LABELS.get(status, status)
        ws.cell(row=row, column=7, value=label)
        ws.cell(row=row, column=8, value=count)
        # Apply status color
        status_fill = STATUS_FILLS.get(status)
        if status_fill:
            ws.cell(row=row, column=7).fill = status_fill
        for col in (7, 8):
            ref = ws.cell(row=style_ref, column=col)
            cell = ws.cell(row=row, column=col)
            copy_cell_style(ref, cell)
            if status_fill and col == 7:
                cell.fill = status_fill
        row += 1

    print(f"  [OK] {cfg['name']}: {len(by_cat)} categories, {len(by_status)} statuses")


def write_detail_sheet(wb, items):
    """Write all expense items to 상세 sheet."""
    cfg = SHEET_CONFIG["detail"]
    ws = wb[cfg["name"]]
    headers = parse_headers(ws, cfg["header_row"])
    data_start = cfg["data_start"]
    style_ref = data_start

    print(f"  [SHEET] {cfg['name']}: headers={list(headers.keys())}")

    # Metadata
    ws.cell(row=2, column=2, value="2025-01")
    ws.cell(row=2, column=5, value="개발팀")
    ws.cell(row=2, column=8, value="자동생성")

    for idx, item in enumerate(items):
        row = data_start + idx
        for header_name, col_idx in headers.items():
            field_key = DETAIL_FIELD_MAP.get(header_name)
            if not field_key:
                continue

            value = item.get(field_key, "")

            # Status label conversion
            if field_key == "policy_status":
                value = STATUS_LABELS.get(value, value)

            cell = ws.cell(row=row, column=col_idx, value=value)
            ref = ws.cell(row=style_ref, column=col_idx)
            copy_cell_style(ref, cell)

            # Status color
            if field_key == "policy_status":
                original_status = item.get("policy_status", "")
                fill = STATUS_FILLS.get(original_status)
                if fill:
                    cell.fill = fill

    print(f"  [OK] {cfg['name']}: {len(items)} items written")


def write_flag_sheet(wb, items):
    """Write flagged items to 플래그 sheet."""
    cfg = SHEET_CONFIG["flag"]
    ws = wb[cfg["name"]]
    headers = parse_headers(ws, cfg["header_row"])
    data_start = cfg["data_start"]
    style_ref = data_start

    flagged = [i for i in items if i.get("policy_status") in ("PENDING", "FLAG", "REVIEW")]
    print(f"  [SHEET] {cfg['name']}: headers={list(headers.keys())}, flagged={len(flagged)}")

    for idx, item in enumerate(flagged):
        row = data_start + idx
        for header_name, col_idx in headers.items():
            field_key = FLAG_FIELD_MAP.get(header_name)
            if not field_key:
                continue

            value = item.get(field_key, "")
            if field_key == "policy_status":
                value = STATUS_LABELS.get(value, value)

            cell = ws.cell(row=row, column=col_idx, value=value)
            ref = ws.cell(row=style_ref, column=col_idx)
            copy_cell_style(ref, cell)

            if field_key == "policy_status":
                original = item.get("policy_status", "")
                fill = STATUS_FILLS.get(original)
                if fill:
                    cell.fill = fill

    print(f"  [OK] {cfg['name']}: {len(flagged)} flagged items written")


# =============================================================================
# Main
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Invoice Classification - Report Generation (Phase 3)"
    )
    parser.add_argument("--judged-data", required=True, help="LLM 판단 완료 데이터")
    parser.add_argument("--output-dir", "-o", required=True, help="출력 디렉토리")
    args = parser.parse_args()

    judged_path = Path(args.judged_data).resolve()
    if not judged_path.exists():
        print(f"[ERROR] File not found: {judged_path}")
        sys.exit(1)

    with open(judged_path, encoding="utf-8") as f:
        data = json.load(f)

    config = load_config()
    items = data.get("items", [])
    summary = data.get("summary", {})

    print(f"\n=== Report Generation (Excel Template) ===\n")

    wb = load_template(config)
    write_summary_sheet(wb, items, summary)
    write_detail_sheet(wb, items)
    write_flag_sheet(wb, items)

    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    # JSON report
    json_path = output_dir / "expense_report.json"
    report = {
        "meta": {
            **data.get("meta", {}),
            "generated_at": datetime.now().isoformat(),
        },
        "items": items,
        "summary": summary,
    }
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"\n  [OK] JSON: {json_path}")

    # Excel report
    xlsx_path = output_dir / "expense_report.xlsx"
    wb.save(xlsx_path)
    print(f"  [OK] Excel: {xlsx_path}")

    print(f"\n[OK] Report generation complete: {output_dir}")


if __name__ == "__main__":
    main()
